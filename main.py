from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from time import sleep
import json
from threading import Thread

from PyQt6 import uic
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QFontDatabase
from PyQt6.QtWidgets import QStyledItemDelegate, QApplication, QMainWindow
from PyQt6.QtWidgets import QTableView, QLabel, QMessageBox, QComboBox, QPushButton
from PyQt6.QtWidgets import QLineEdit, QTextEdit, QProgressBar, QFileDialog, QInputDialog
import sys
import os
import subprocess

from openpyxl import Workbook

Ui_MainWindow, QtBaseClass = uic.loadUiType('window.ui')

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)

        self.workbook = Workbook()
        self.sheet = self.workbook.active

        self.sheet.cell(row=1, column=1).value = 'Backlink'
        self.sheet.cell(row=1, column=2).value = 'Result'

        self.start_row = 2

        self.mode = 'Normal'

        self.running = False
        self.done = False

        self.names = []
        self.names_content = ""
        self.emails = []
        self.emails_content = ""
        self.messages = []
        self.messages_content = ""
        self.backlinks = []
        self.backlins_content = ""
        self.website_link = ''
        self.count_names = 0
        self.count_emails = 0
        self.count_messages = 0
        self.count_backlinks = 0

        self.current_count = 0
        self.current_count_table = 0
        self.current_count_success = 0
        self.current_count_failed = 0

        self.upload_count = 0
        self.add_website_success = False
        self.setting_success = False

        self.backlink_success = []
        self.backlink_failed = []

        self.select_mode = self.findChild(QComboBox, 'select_mode')
        self.select_mode.setCurrentIndex(0)
        self.select_mode.currentIndexChanged.connect(self.select_mode_index_changed)
        self.select_speed = self.findChild(QComboBox, 'select_speed')
        self.select_speed.setCurrentIndex(0)

        self.button_names = self.findChild(QPushButton, 'button_names')
        self.button_names.clicked.connect(self.names_button_clicked)
        self.button_emails = self.findChild(QPushButton, 'button_emails')
        self.button_emails.clicked.connect(self.emails_button_clicked)
        self.button_messages = self.findChild(QPushButton, 'button_messages')
        self.button_messages.clicked.connect(self.messages_button_clicked)
        self.button_backlinks = self.findChild(QPushButton, 'button_backlinks')
        self.button_backlinks.clicked.connect(self.backlinks_button_clicked)
        self.button_clear = self.findChild(QPushButton, 'button_clear')
        self.button_clear.clicked.connect(self.clear_button_clicked)
        self.button_start_stop = self.findChild(QPushButton, 'button_start_stop')
        self.button_start_stop.clicked.connect(self.start_stop_button_clicked)
        self.button_start_stop.setEnabled(self.setting_success)
        self.button_open_success = self.findChild(QPushButton, 'button_open_success')
        self.button_open_success.clicked.connect(self.open_success_button_clicked)
        self.button_open_failed = self.findChild(QPushButton, 'button_open_failed')
        self.button_open_failed.clicked.connect(self.open_failed_button_clicked)
        self.button_add_website = self.findChild(QPushButton, 'button_add_website')
        self.button_add_website.clicked.connect(self.add_website_button_clicked)

        self.label_count_names = self.findChild(QLabel, 'count_names')
        self.label_count_names.setText(str(self.count_names))
        self.label_count_emails = self.findChild(QLabel, 'count_emails')
        self.label_count_emails.setText(str(self.count_emails))
        self.label_count_messages = self.findChild(QLabel, 'count_messages')
        self.label_count_messages.setText(str(self.count_messages))
        self.label_count_backlinks = self.findChild(QLabel, 'count_backlinks')
        self.label_count_backlinks.setText(str(self.count_backlinks))
        self.label_count_success = self.findChild(QLabel, 'count_success')
        self.label_count_success.setText(str(self.current_count_success))
        self.label_count_failed = self.findChild(QLabel, 'count_failed')
        self.label_count_failed.setText(str(self.current_count_failed))

        self.list_names = self.findChild(QTextEdit, 'list_names')
        self.list_names.setPlaceholderText("name list")
        self.list_emails = self.findChild(QTextEdit, 'list_emails')
        self.list_emails.setPlaceholderText("email list")
        self.list_messages = self.findChild(QTextEdit, 'list_messages')
        self.list_messages.setPlaceholderText("message list")
        self.list_backlinks = self.findChild(QTextEdit, 'list_backlinks')
        self.list_backlinks.setPlaceholderText("backlink list")

        self.add_website = self.findChild(QLineEdit, 'add_website')
        self.add_website.setPlaceholderText("Enter your website url")
        self.add_website.setText(self.website_link)

        self.table_view = self.findChild(QTableView, 'tableView')
        self.model = QStandardItemModel(0, 6, self)
        self.table_view.setModel(self.model)
        self.add_table({})
        
        self.progressBar = self.findChild(QProgressBar, 'progressBar')
        self.progressBar.setValue(0)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_timer)

        # service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
        # options = Options()
        # options.add_experimental_option("debuggerAddress", "127.0.0.1:9015")
        # self.driver = webdriver.Chrome(service=service, options=options)
    
    def start_timer(self):
        self.timer.start(500)
    
    def stop_timer(self):
        self.timer.stop()

    def update_timer(self):
        if self.done == True:
            self.stop_timer()
            result = QMessageBox.information(self, "Notification", "Done!", QMessageBox.StandardButton.Ok)
            if result == QMessageBox.StandardButton.Ok:
                pass

    def select_mode_index_changed(self):
        self.mode = self.select_mode.currentText()
        print(self.mode)
   
    def names_button_clicked(self):
        self.open_file('name')

    def emails_button_clicked(self):
        self.open_file('email')

    def messages_button_clicked(self):
        self.open_file('message')

    def backlinks_button_clicked(self):
        self.open_file('backlink')

    def open_file(self, button_name):
        file_dialog = QFileDialog()
        file_dialog.setWindowTitle("Open Name File")

        # Set the file mode to allow selecting only existing files
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)

        # Set the file filter to allow only certain file types
        file_dialog.setNameFilters(["Text Files (*.txt)"])

        # Open the dialog and wait for user selection
        content = ''
        if file_dialog.exec() == QFileDialog.DialogCode.Accepted:
            selected_files = file_dialog.selectedFiles()
            for selected_file in selected_files:
                with open(selected_file, 'r') as file:
                    content += file.read()

            self.upload_success = False
            self.setting_success = False
            self.button_start_stop.setEnabled(self.setting_success)

            if button_name == "name":
                self.list_names.setPlainText(content)
                self.names = content.split('\n')
                self.count_names = len(self.names)
                self.label_count_names.setText(str(self.count_names))

            if button_name == "email":
                if "@" not in content:
                    content == ''
                self.list_emails.setPlainText(content)
                self.emails = content.split('\n')
                self.count_emails = len(self.emails)
                self.label_count_emails.setText(str(self.count_emails))

            if button_name == "message":
                self.list_messages.setPlainText(content.replace('\n', '\n\n'))
                self.messages = content.split('\n')
                self.count_messages = len(self.messages)
                self.label_count_messages.setText(str(self.count_messages))

            if button_name == "backlink":
                if "http" not in content:
                    content == ''
                self.list_backlinks.setPlainText(content)
                self.backlinks = content.split('\n')
                self.count_backlinks = len(self.backlinks)
                self.label_count_backlinks.setText(str(self.count_backlinks))

            if self.count_names > 0 and self.count_emails > 0 and self.count_messages > 0 and self.count_backlinks > 0:
                self.upload_success = True
            else:
                self.upload_success = False
            self.setting_success = self.upload_success and self.add_website_success
            self.button_start_stop.setEnabled(self.setting_success)
    
    def clear_button_clicked(self):
        result = QMessageBox.information(self, "Question", "Do you want to really clear?", QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if result == QMessageBox.StandardButton.Yes:  
            self.model.clear()
            self.progressBar.setValue(0)
            self.select_mode.setEnabled(True)
            self.select_speed.setEnabled(True)
            self.button_emails.setEnabled(True)
            self.button_names.setEnabled(True)
            self.button_messages.setEnabled(True)
            self.button_backlinks.setEnabled(True)
            self.button_add_website.setEnabled(True)

    def open_success_button_clicked(self):
        absolute_path = os.path.abspath("backlink_success.json")
        if os.path.exists(os.path.abspath(absolute_path)):
            subprocess.run(["notepad.exe", absolute_path])
        else:
            result = QMessageBox.information(self, "Warning", "File does not exist", QMessageBox.StandardButton.Close)
            if result == QMessageBox.StandardButton.Close:
                return

    def open_failed_button_clicked(self):
        absolute_path = os.path.abspath("backlink_failed.txt")
        if os.path.exists(os.path.abspath(absolute_path)):
            subprocess.run(["notepad.exe", absolute_path])
        else:
            result = QMessageBox.information(self, "Warning", "File does not exist", QMessageBox.StandardButton.Close)
            if result == QMessageBox.StandardButton.Close:
                return

    def add_website_button_clicked(self):
        text, ok = QInputDialog.getText(self, "Input Dialog", "Enter your website url:")
        if ok:
            if ("http://" not in text and "https://" not in text) or "." not in text:
                result = QMessageBox.information(self, "Warning", "Wrong url", QMessageBox.StandardButton.Close)
                if result == QMessageBox.StandardButton.Close:
                    self.add_website.setText('')
                    self.add_website_success = False
                    return    
            self.website_link = text
            self.add_website.setText(text)
            self.add_website_success = True
            self.setting_success = self.add_website_success and self.upload_success
            self.button_start_stop.setEnabled(self.setting_success)

    def add_table(self, result):
        try:
            self.model.setHorizontalHeaderLabels(["No", "Backlink", "Name", "Email", "Message", "Status"])
            if len(result) != 0:
                self.current_count_table += 1
                if self.count_backlinks != 0:
                    self.progressBar.setValue(int(self.current_count_table * 100 / self.count_backlinks))
                else:
                    self.progressBar.setValue(0)

                items = [str(self.current_count_table), result["url"], result["name"], result["email"], result["message"], result["status"]]
                if len(result) > 0:
                    row_count = self.model.rowCount()
                    self.model.insertRow(row_count)
                    column = 0
                    for value in items:
                        item = QStandardItem(str(value))
                        self.model.setItem(row_count, column, item)
                        column += 1
            [self.table_view.setColumnWidth(index, value) for index, value in enumerate([81, 500, 150, 150, 300, 100])]
        except:
            pass
        
    def start_stop_button_clicked(self):
        if self.running:
            self.running = False
            self.select_mode.setEnabled(True)
            self.select_speed.setEnabled(True)
            self.button_emails.setEnabled(True)
            self.button_names.setEnabled(True)
            self.button_messages.setEnabled(True)
            self.button_backlinks.setEnabled(True)
            self.button_add_website.setEnabled(True)

            with open('backlink_success.json', 'w') as file:
                json.dump(self.backlink_success, file)

            with open('backlink_failed.txt', 'w') as file:
                json.dump(self.backlink_failed, file)
            
            self.workbook.save('output.xlsx')

        else:
            self.running = True
            self.select_mode.setEnabled(False)
            self.select_speed.setEnabled(False)
            self.button_emails.setEnabled(False)
            self.button_names.setEnabled(False)
            self.button_messages.setEnabled(False)
            self.button_backlinks.setEnabled(False)
            self.button_add_website.setEnabled(False)

            thread = Thread(target=self.main)
            thread.start()
            self.start_timer()

    def submit(self, backlink, message, name, email, mywebsite):
        status = ''
        if backlink == '':
            self.backlink_failed.append(backlink)
            status = "Failed"
        else:
            try:
                driver = webdriver.Chrome()
                driver.get(backlink)
                if self.running:
                    try:
                        driver.find_element(By.ID, 'comment').send_keys(message)
                        driver.find_element(By.ID, 'author').send_keys(name)
                        driver.find_element(By.ID, 'email').send_keys(email)
                        driver.find_element(By.ID, 'url').send_keys(mywebsite)
                        submit_button = driver.find_element(By.ID, 'submit')
                        driver.execute_script("arguments[0].click();", submit_button)

                        self.backlink_success.append({"url" : backlink, "name" : name, "email": email, "message" : message})
                        status = "Success"
                    except:
                        self.backlink_failed.append(backlink)
                        status = "Failed"
            except:
                self.backlink_failed.append(backlink)
                status = "Failed"

        self.add_table({"url" : backlink, "name" : name, "email": email, "message" : message, "status": status})
        self.current_count_failed += 1
        self.label_count_failed.setText(str(self.current_count_failed))
        self.sheet.cell(row=self.start_row, column=1).value = backlink
        self.sheet.cell(row=self.start_row, column=2).value = status
        self.start_row += 1

    def main(self):
        speed = int(self.select_speed.currentText())
        threads = []
        if speed > self.count_backlinks:
            speed = self.count_backlinks
        for i in range(speed):
            message = self.messages[self.current_count % self.count_messages]
            email = self.emails[self.current_count % self.count_emails]
            name = self.names[self.current_count % self.count_names]
            backlink = self.backlinks[self.current_count]
            self.current_count += 1
            thread = Thread(target=self.submit, args=(backlink, message, name, email, self.website_link))
            thread.start()
            threads.append(thread)

        while self.running and self.current_count < self.count_backlinks:
            for index, thread in enumerate(threads):
                if not thread.is_alive():
                    message = self.messages[self.current_count % self.count_messages]
                    email = self.emails[self.current_count % self.count_emails]
                    name = self.names[self.current_count % self.count_names]
                    backlink = self.backlinks[self.current_count]
                    self.current_count += 1
                    thread = Thread(target=self.submit, args=(backlink, message, name, email, self.website_link))
                    threads[index] = thread
                    thread.start()
                if self.current_count > self.count_backlinks:
                    break
            sleep(0.1)

        while self.current_count_success + self.current_count_failed < self.count_backlinks and self.running:
            print(f'{self.count_backlinks}({self.current_count_success + self.current_count_failed})')
            sleep(1)

        if self.running:
            self.running = False
            self.upload_success = False
            self.setting_success = False   

            with open('backlink_success.json', 'w') as file:
                json.dump(self.backlink_success, file)

            with open('backlink_failed.txt', 'w') as file:
                json.dump(self.backlink_failed, file)
            
            self.workbook.save('output.xlsx')
            self.done = True
            print('Done!')

if __name__ == "__main__":
    app = QApplication([])
    window = MainWindow()
    window.show()
    sys.exit(app.exec())